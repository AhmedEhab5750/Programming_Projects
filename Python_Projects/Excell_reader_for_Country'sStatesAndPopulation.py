# Program: First application. The app about that it can print an Excel file .xlsx in the same folder with
# the code file by writing the file name in first choice you start with writing 1
# after that you write the file name like if you got Egypt.xlsx in the same folder write Egypt, and now
# you can do one of two functions first function: print the excell file made of 2 columns about states and population
# and at the same time it also prints the total population after the table.
# Second function is to print the state with the biggest population and the one with the smallest.
# Version: 4.0
# Date: 3/3/2024
import openpyxl


def check_input(choice):
    while True:
        # Try to convert the input to an integer and return its Value
        if choice.lower() == 'exit':
            return 4
        else:
            try:
                value = int(choice)
                if 0 < value < 5:  # Check if input is between 1-7
                    return value
                else:
                    print("Invalid Input, Input a value between 1-4 or type \'exit\'")
            # If the input is not an integer, print a message and try again
            except Exception as e:
                print(f"{e}.\n'{choice}' is an invalid choice. Please enter a number from 1 to 4")
                return


def load_file_1():
    try:
        country = input("Please enter the country name: ")
        # Try loading the existing workbook
        work_book = openpyxl.load_workbook(f'{country}.xlsx')
        return work_book
    except FileNotFoundError:
        print("The file does not exist.")


def display_all_population_2(current_work_book):
    try:
        for sheet_name in current_work_book.sheetnames:
            total_population = 0

            sheet = current_work_book[sheet_name]
            print("+-------------------------+-------------------------+")
            print("|State                    |               Population|")
            print("+-------------------------|-------------------------+")
            for row in sheet.iter_rows(min_row=1, values_only=True):
                state, population = row[0], row[1]
                print(f"|{state:<25}|{population:>25}|")
                print("+-------------------------|-------------------------+")
                total_population += population
            print(f"\nTotal Population: {total_population}")

    except:
        print("Choose a country that got an .xlsx file first")
        return


def display_highest_lowest_population_3(current_work_book):
    try:
        for sheet_name in current_work_book.sheetnames:
            sheet = current_work_book[sheet_name]
            max_population = 0
            min_population = float('inf')
            state_max_population = ""
            state_min_population = ""
            for row in sheet.iter_rows(min_row=1, values_only=True):
                state, population = row[0], row[1]
                if population > max_population:
                    max_population = population
                    state_max_population = state
                if population < min_population:
                    min_population = population
                    state_min_population = state
            print(f"State/Province/Governorate with Highest Population: {state_max_population} - {max_population}")
            print(f"State/Province/Governorate with Lowest Population: {state_min_population} - {min_population}")
    except:
        print("Choose a country that got an .xlsx file first")
        return


current_wb = None
while True:
    try:
        print("This app about statistics of different countries")
        print("1) Choose Country")
        print("2) List each state and its population, and Total of population of country")
        print("3) Print States of highest and lowest Population")
        print("4) Exit")
        choice = input("Enter your choice: ")
        choice = check_input(choice)
        if choice == 1:
            current_wb = load_file_1()
        elif choice == 2:
            display_all_population_2(current_wb)
        elif choice == 3:
            display_highest_lowest_population_3(current_wb)
        elif choice == 4:
            print("Thank you for using our program.")
            break
    except Exception as e:
        print(f"{e}.\n'{choice}' is an invalid choice. Please enter a number from 1 to 4")